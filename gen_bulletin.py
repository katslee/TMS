import openpyxl
from datetime import datetime, date, timedelta
import os
from subprocess import call
import glob
import os
import shutil
import gen_ordering
import unicodedata
from datetime import datetime, timedelta
from subprocess import call

import openpyxl

import gen_ordering

crlf = chr(13) + chr(10)
lf = chr(10)

# Mac setting
#text_output = "/Users/Kats/Documents/TickerManagementSystem/Python/TextBulletin/"
#graphic_output = "/Users/Kats/Documents/TickerManagementSystem/Python/GraphicBulletin/"
#working = "/Users/Kats/Documents/TickerManagementSystem/Python/working/"
#pythonfolder = "/Users/Kats/Documents/TickerManagementSystem/Python/Python/"
#updatefolder = "/Users/Kats/Documents/TickerManagementSystem/Python/update/"
#errorfolder = "/Users/Kats/Documents/TickerManagementSystem/Python/error/"
#convertedfolder = "/Users/Kats/Documents/TickerManagementSystem/Python/converted/"

# AMS UAT Server
text_output = "/data1/TMS/phrase1/network/export/result/TextBulletin/"
graphic_output = "/data1/TMS/phrase1/network/export/result/GraphicBulletin/"
output = "/data1/TMS/phrase1/network/export/result/"
graphicengine1 = "/data1/TMS/phrase1/network/graphicengine1/result/GraphicBulletin/"
graphicengine2 = "/data1/TMS/phrase1/local/graphicengine2/result/GraphicBulletin/"
textengine1 = "/data1/TMS/phrase1/network/graphicengine1/result/TextBulletin/"
textengine2 = "/data1/TMS/phrase1/local/graphicengine2/result/TextBulletin/"
working = "/data1/TMS/phrase1/working/"
pythonfolder = "/data1/TMS/phrase1/python/"
updatefolder = "/data1/TMS/phrase1/update/"
errorfolder = "/data1/TMS/phrase1/network/export/error/"
convertedfolder = "/data1/TMS/phrase1/network/export/converted/"

text_bulletin_filename = "L-Title.txt"

def unilen(line):
    cnt = 0
    for c in line:
        cat = unicodedata.category(c)
        if cat == 'Lo':
            cnt += 1
        elif cat[0] == 'P':
            cnt += 0.5
        else:
            cnt += 0.5
    cnt = int(cnt + 0.5)
    return cnt

def read_excel(filename):
    global errorfolder

    #folder = datetime.now().strftime("%d%m%Y%H%M%S") + "/"
    folder = datetime.now().strftime("%Y%m%d%H%M") + "/"
    #os.mkdir(output + folder)
    files = glob.glob(working + "*")
    for f in files:
            os.remove(f)
    errorfolder = errorfolder + folder
    if not os.path.exists(errorfolder):
        os.mkdir(errorfolder)
    with open(errorfolder + 'error_' + os.path.basename(filename) + '.txt', "w") as errfile:
        errfile.writelines(datetime.now().strftime("%Y%m%d%H%M") + "\r\n")
    error = False
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.worksheets[0]
    current_time = datetime.now()

    script = working + "genbulletin.sh"
    with open(script, "w") as f:
        f.writelines("# " + datetime.now().strftime("%d/%m/%Y %H:%M:%S") + "\n")
        f.writelines("cd " + working + "\n")

    row = 2
    snlist = []
    while ws.cell(row=row,column=1).value != None:
        bulletinType = str.lower(ws.cell(row=row, column=1).value)
        if bulletinType == "text bulletin":
            bulletinType = "T"
        elif bulletinType == "graphic bulletin":
            bulletinType = "G"
        else:
            with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                errfile.writelines(sn + " - Ticker type undefined (" + bulletinType + ")." + "\r\n")
            error = True
        sn = str(ws.cell(row=row, column=11).value)
        snlist.append(sn)
        title = ws.cell(row=row, column=12).value

        if title.count(lf) > 1:
            with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                errfile.writelines(sn + " - title line exceed (" + str(title.count(lf)) + ")." + "\r\n")
            error = True
        for line in title.splitlines():
            if (bulletinType == "G" and unilen(line) > 11) or (bulletinType == "T" and unilen(line) > 14):
                with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                    errfile.writelines(sn +  " - title words exceed (" + str(unilen(line)) + ")." + "\r\n")
                error = True

        content = ws.cell(row=row, column=13).value
        if (bulletinType == "G" and content.count(lf) > 2) or (bulletinType == "T" and content.count(lf) > 8):
            with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                errfile.writelines(sn + " - content line exceed (" + str(content.count(lf)) + ")." + "\r\n")
            error = True
        for index, line in enumerate(content.splitlines()):
            if (bulletinType == "G" and unilen(line) > 7) or (bulletinType == "T" and unilen(line) > 7):
                with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                    errfile.writelines(sn + " - content words exceed (" + str(index + 1) + ":" + str(unilen(line)) + ")." + "\r\n")
                error = True
        content = content.replace(lf, crlf)

        footer = ws.cell(row=row, column=14).value
        if (bulletinType == "G" and footer == None):
            footer = " "
        else:
            if (bulletinType == "T" and footer != None):
                with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                    errfile.writelines(sn + " - should not has footer." + "\r\n")
                error = True

        if (bulletinType =="G" and footer.count(lf) > 0):
            with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                errfile.writelines(sn + " - footer line exceed ("  + str(footer.count(lf)) + ")." + "\r\n")
            error = True

        if bulletinType == "G":
            if footer != None:
                for line in footer.splitlines():
                    if unilen(line) > 10:
                        with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                            errfile.writelines(sn + " - footer words exceed (" + str(unilen(line)) + ")." + "\r\n")
                        error = True
            else:
                with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                    errfile.writelines(sn + " - No footer." + "\r\n")
                error = True

        qrcode = ws.cell(row=row, column=15).value
        if (bulletinType == "G" and qrcode == None):
            with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                errfile.writelines(sn + " - Graphic Bulletin, no QRCode URL." + "\r\n")
            error = True
        if (bulletinType == "T" and qrcode != None):
            with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                errfile.writelines(sn + " - Text Bulletin, has QRCode URL." + "\r\n")
            error = True

        txdate = ws.cell(row=row, column=5).value
        enddate = ws.cell(row=row, column=7).value
        if enddate == None:
            enddate = txdate + timedelta(days=14)
        etime = ws.cell(row=row, column=8).value
        if etime == None:
            endtime = timedelta(hours=23, minutes=59)
        else:
            endtime = timedelta(hours=int(etime / 100), minutes=etime % 100)
        EndTime = endtime + enddate

        fname = working + "title" + sn + ".txt"
        with open(fname, "w") as f:
            f.writelines(title)
        f.close()

        fname = working  + "content" + sn + ".txt"
        with open(fname, "w") as f:
            f.writelines(content)
        f.close()

        fname = working + "footer" + sn + ".txt"
        with open(fname, "w") as f:
            f.writelines(str(footer))
        f.close()

        with open(script, "a") as f:
            # Add leading 0 is month, day, hour & minute is single digit
            if EndTime.month < 10:
                dcode = "0" + str(EndTime.month)
            else:
                dcode = str(EndTime.month)
            if EndTime.day < 10:
                dcode = dcode + "0" + str(EndTime.day)
            else:
                dcode = dcode + str(EndTime.day)
            if EndTime.hour < 10:
                tcode = "0" + str(EndTime.hour)
            else:
                tcode = str(EndTime.hour)
            if EndTime.minute < 10:
                tcode = tcode + "0" + str(EndTime.minute)
            else:
                tcode = tcode + str(EndTime.minute)
            if bulletinType == "G":
                if EndTime >= current_time and qrcode != None:
                    f.writelines(pythonfolder + "upper_image_billboard.sh " + str(sn) + " " + chr(34) + qrcode + chr(34) + " " + dcode + " " + tcode + "\n")
            else:
                with open(working + gen_ordering.fname(sn, EndTime, "T"), "w") as f:
                    f.writelines(title + "\r\n")
                    f.writelines(content + "\r\n")
                    f.writelines("\r\n")

            row += 1
        f.close()

# Check any SN is duplicated in Excel
    newSNlist = []
    dupSNlist = []
    for s in snlist:
        if s not in newSNlist:
            newSNlist.append(s)
        else:
            dupSNlist.append(s)
    if len(dupSNlist) > 0:
        with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
            for sn in dupSNlist:
                errfile.writelines(sn + " - Duplicated." + "\r\n")
        error = True

    os.chmod(script,0o755)
    for root, dirs, files in os.walk(working):
        for file in files:
            os.chmod(os.path.join(root, file), 0o777)
    if not error:
        # Delete all files in updatefolder
        files = glob.glob(updatefolder + "*")
        for f in files:
            os.remove(f)
        # Copy the excel file to update & working folder
        path, fname = os.path.split(filename)
        ufolder = os.path.join(updatefolder,fname)
        try:
            shutil.copy2(filename, convertedfolder)
        except:
            with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                errfile.writelines("Excel file cannot move in ingest folder." + "\r\n")

        try:
            shutil.move(filename, ufolder)
        except:
            with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                errfile.writelines("Excel file cannot move to update folder." + "\r\n")

        rcode = call(script, shell=True)

        # Move all files in output folder to old folder
        files = glob.iglob(os.path.join(graphic_output, "*.jpg"))
        for f in files:
            if os.path.isfile(f):
                path, fname = os.path.split(f)
                try:
                    shutil.move(f, graphic_output + "old/" + fname)
                except:
                    with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                        errfile.writelines("Cannot move graphic from output folder to old folder." + "\r\n")

        files = glob.iglob(os.path.join(graphicengine1, "*.jpg"))
        for f in files:
            if os.path.isfile(f):
                path, fname = os.path.split(f)
                try:
                    shutil.move(f, graphicengine1 + "old/" + fname)
                except:
                    with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                        errfile.writelines("Cannot move graphic from graphengine1 to old folder." + "\r\n")

        files = glob.iglob(os.path.join(graphicengine2, "*.jpg"))
        for f in files:
            if os.path.isfile(f):
                path, fname = os.path.split(f)
                try:
                    shutil.move(f, graphicengine2 + "old/" + fname)
                except:
                    with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                        errfile.writelines("Cannot move graphic from graphengine2 to old folder." + "\r\n")


        # Copy jpg from working folder to result folder
        files = glob.iglob(os.path.join(working, "*.jpg"))
        for f in files:
            if os.path.isfile(f):
                # Copy to update folder
                shutil.copy2(f, updatefolder)
                try:
                    shutil.copy2(f, graphic_output)
                except:
                    with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                        errfile.writelines("Cannot copy graphic from working to graphic folder." + "\r\n")


                # copy to Graphicengine1
                try:
                    shutil.copy2(f, graphicengine1)
                except:
                    with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                        errfile.writelines("Cannot copy graphic from working to graphicengine1 folder." + "\r\n")


                # copy to Graphicengine2
                try:
                    shutil.copy2(f, graphicengine2)
                except:
                    with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
                        errfile.writelines("Cannot copy graphic from working to graphicengine2 folder." + "\r\n")

        # Copy text from working to update folders
        files = glob.iglob(os.path.join(working, "*.txt"))
        for f in files:
            if os.path.isfile(f):
                shutil.copy2(f, updatefolder)

        # Generate ordering file
#        gen_ordering.gen_order(ufolder, graphic_output, text_output)
        # Copy to update folder
#        shutil.copy2(graphic_output + "gb_order.txt", updatefolder)
#        shutil.copy2(text_output + "L-Title.txt", updatefolder)

        gen_ordering.gen_order(ufolder,updatefolder,updatefolder)
        shutil.copy2(updatefolder + "gb_order.txt", graphic_output)
        shutil.copy2(updatefolder + "gb_order.txt", graphicengine1)
        shutil.copy2(updatefolder + "gb_order.txt", graphicengine2)
        shutil.copy2(updatefolder + "L-Title.txt", text_output)
        shutil.copy2(updatefolder + "L-Title.txt", textengine1)
        shutil.copy2(updatefolder + "L-Title.txt", textengine2)

        with open(errorfolder + "error_" + os.path.basename(filename) + ".txt", "a") as errfile:
            errfile.writelines(filename + " - Success." + "\r\n")
    else:
        shutil.move(filename, errorfolder)

    errfile.close()



