import openpyxl
from datetime import datetime, date, timedelta
import glob
import os


# AMS UAT Server
#watch = "/data1/TMS/phrase1/network/export/"
#update = "/data1/TMS/phrase1/update/"

# MacOS Development
watch = "/Users/Kats/Documents/TickerManagementSystem/Python/watch/"
update = "/Users/Kats/Documents/TickerManagementSystem/Python/update/"
def takefrequency(elem):
    return elem[1]

def takepriority(elem):
    return elem[2]

def takesn(elem):
    return elem[0]

def fname(sn, etime, btype):
    if etime.month < 10:
        dcode = "0" + str(etime.month)
    else:
        dcode = str(etime.month)
    if etime.day < 10:
        dcode = dcode + "0" + str(etime.day)
    else:
        dcode = dcode + str(etime.day)
    if etime.hour < 10:
        tcode = "0" + str(etime.hour)
    else:
        tcode = str(etime.hour)
    if etime.minute < 10:
        tcode = tcode + "0" + str(etime.minute)
    else:
        tcode = tcode + str(etime.minute)
    if btype == "G":
        fn = sn + " d"  + dcode + " " + tcode + ".jpg"
    else:
        fn = sn + " d"  + dcode + " " + tcode + ".txt"
    return fn

def remove_dup(bulletins):
    last_bulletin = []
    new_bulletins = []
    for b in bulletins:
        if last_bulletin != b:
            new_bulletins.append(b)
        last_bulletin = b
    return new_bulletins

def reorder(bulletins):
    last = len(bulletins) - 1
    if bulletins[last] == bulletins[last - 1]:
        b = bulletins[last]
        l = last - 1
        for i in reversed(range(l)):
            if (bulletins[i] != b) and (bulletins[i - 1] != b):
                bulletins.insert(i,b)
                break
        del bulletins[-1]
    return bulletins

def gen_order(filename, gfolder, tfolder):
# Bulletin [sn, bulletinType, frequency, priority, channel, TXTime, EndTime, Filename]

    wb = openpyxl.load_workbook(filename,data_only=True)
    ws = wb.worksheets[0]

    row = 2
    g_bulletins = []
    t_bulletins = []
    while ws.cell(row=row,column=1).value != None:
        sn = str(ws.cell(row=row, column=11).value)
        bulletinType = str.lower(ws.cell(row=row, column=1).value)
        if bulletinType == "text bulletin":
            bulletinType = "T"
        elif bulletinType == "graphic bulletin":
            bulletinType = "G"

        frequency = ws.cell(row=row, column=2).value
        priority = ws.cell(row=row, column=3).value
        channel = ws.cell(row=row, column=4).value
        txdate = ws.cell(row=row, column=5).value
        txtime = ws.cell(row=row, column=6).value
        if txtime == None:
            txtime = timedelta(hours=0,minutes=0)
        TXTime = txdate + txtime
        enddate = ws.cell(row=row, column=7).value
        if enddate == None:
            enddate  = txdate + timedelta(days=14)
        etime = ws.cell(row=row, column=8).value
        if etime == None:
            endtime = timedelta(hours=23,minutes=59)
        else:
            endtime = timedelta(hours=int(etime/100),minutes=etime % 100)
        EndTime = endtime + enddate
        bulletin = []
        bulletin.append(sn)
        bulletin.append(frequency)
        bulletin.append(priority)
        bulletin.append(channel)
        bulletin.append(TXTime)
        bulletin.append(EndTime)
        bulletin.append(fname(sn,EndTime,bulletinType))
        if bulletinType == "T":
            t_bulletins.append(bulletin)
        else:
            g_bulletins.append(bulletin)
        row += 1

    wb.close()

# Graphic Bulletin Order (SN / Priority / Frequency)
    g_bulletins.sort(key=takesn, reverse=True)
    g_bulletins.sort(key=takefrequency,reverse=True)
    g_bulletins.sort(key=takepriority, reverse=False)

    g_order = []

# Round 1 allocation until no append SN in a loop
    finish = False
    current_time = datetime.now()
    while not finish:
        finish = True
        for bulletin in g_bulletins:
            if (bulletin[1] > 0) and (bulletin[4] < current_time) and (current_time <= bulletin[5]):
                g_order.append(bulletin[6])
                if len(g_order) > 2:
                    reorder(g_order)
                bulletin[1] -= 1
                finish = False

# Next Round if anyone Frequency > 0
    finish = False
    while not finish:
        finish = True
        for bulletin in g_bulletins:
            if (bulletin[1] > 0) and (bulletin[4] < current_time) and (current_time <= bulletin[5]):
                g_order.append(bulletin[6])
                reorder(g_order)
                bulletin[1] -= 1
                finish = False


# Text Bulletin Order (SN > Priority > Frequency)
    t_bulletins.sort(key=takesn, reverse=True)
    t_bulletins.sort(key=takefrequency, reverse=True)
    t_bulletins.sort(key=takepriority, reverse=False)

    t_order = []
    finish = False
    while not finish:
        finish = True
        for bulletin in t_bulletins:
            current_time = datetime.now()
            if (bulletin[1] > 0) and (bulletin[4] < current_time) and (current_time <= bulletin[5]):
                #t_order.append(bulletin[6] + "," + str(bulletin[2]))
                t_order.append(bulletin[6])
                bulletin[1] -= 1
                finish = False

# Next Round if anyone Frequency > 0
    finish = False
    while not finish:
        finish = True
        for bulletin in t_bulletins:
            if (bulletin[1] > 0) and (bulletin[4] < current_time) and (current_time <= bulletin[5]):
                t_order.append(bulletin[6])
                reorder(t_order)
                bulletin[1] -= 1
                finish = False

#    g_order = remove_dup(g_order)
#    t_order = remove_dup(t_order)

    g_order_file = gfolder + 'gb_order.txt'
    with open(g_order_file, "w") as f:
        for g in g_order:
            f.writelines(g + "\r\n")
        f.close()

    t_order_file = tfolder + 'L-Title.txt'
    with open(t_order_file, "w") as f:
        for t in t_order:
            # read text bulletin from update folder and write to L-Title
            with open(update + t, "r") as tfile:
                tcontent = tfile.readlines()
                for t in tcontent:
                    t = t.replace("\n","\r\n")
                    f.writelines(t)
            tfile.close()
        f.close()


